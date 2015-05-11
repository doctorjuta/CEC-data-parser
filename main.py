from bs4 import BeautifulSoup
from urllib.request import urlopen
from openpyxl import Workbook
from xlsxdata import get_dep_datas, count_dep_list

start_url = "http://www.cvk.gov.ua/pls/vnd2012/wp401?PT001F01=900"
base_url = "http://www.cvk.gov.ua/pls/vnd2012/"
content = urlopen(start_url).read()
soup = BeautifulSoup(content)

workbook = Workbook()
worksheet_list = workbook.active
worksheet_list.title = 'Депутати у списках'
worksheet_vo = workbook.create_sheet(title='Депутати в округах')
worksheet_part_list = workbook.create_sheet(title='Партії зі списками')
row = [1,1,1] # Перше значення - позиція запису на лиску списку, друга - на листку в округах, третє - загальний порядок

tables = soup.find_all('table', 't2')
linksraw = tables[0].find_all('a')
links = [start_url]

# Збираємо усі посилання на алфавітні сторінки з депутататами
for link in linksraw:
    links.append(base_url+link['href'].replace("¤", "&curren"))

# Записуємо дані депутата з кожного посилання у xlsx файл
for link in links:
    row = get_dep_datas(link, worksheet_list, worksheet_vo, row, base_url, BeautifulSoup, urlopen)

# Перераховуємо кількість унікальних партій зі списками депутатів і рахуємо кількість депутатів у списках кожної
part_cnt = count_dep_list(worksheet_part_list, worksheet_list)
print(str(part_cnt)+" parties wrote in document")

"""
rese = ""
f = open('text.txt', 'w')
f.write(rese)
f.close()
"""

workbook.save("res.xlsx")